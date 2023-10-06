from datetime import datetime, timedelta
from pprint import pprint

from django.db.models import Count
from django.http import HttpResponse

from openpyxl import Workbook
from robots.models import Robot


def download_robot_report(request):
    end_date = datetime.now()
    start_date = end_date - timedelta(days=7)
    robot_data = Robot.objects.filter(
        created__gte=start_date, created__lte=end_date
    ).values('model', 'version').annotate(count=Count('id'))
    report_data = {}
    pprint(robot_data)
    for robot in robot_data:
        model = robot['model']
        version = robot['version']
        count = robot['count']

        if model not in report_data:
            report_data[model] = {}

        report_data[model][version] = count
    pprint(report_data)
    return generate_excel_file(report_data)


def generate_excel_file(robot_data):
    wb = Workbook()
    for model in robot_data:
        ws = wb.create_sheet(title=model)

        ws.append(["Модель", "Версия", "Количество за неделю"])
        for version, count in robot_data[model].items():
            ws.append([model, version, count])

    del wb[wb.sheetnames[0]]

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=robot_export.xlsx'

    wb.save(response)

    return response